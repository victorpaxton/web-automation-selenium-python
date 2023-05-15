import unittest

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from selenium.common.exceptions import NoSuchElementException

from openpyxl import load_workbook


class Utils:
    def login(driver):
        wait = WebDriverWait(driver, 2)
        username = wait.until(EC.element_to_be_clickable((By.ID, "username")))
        username.clear()
        username.send_keys("teacher")
        password = driver.find_element(by=By.ID, value="password")
        password.clear()
        password.send_keys("moodle")
        password.send_keys(Keys.RETURN)

    def read_data_from_excel(filename, sheet):
        datalist = []
        wb = load_workbook(filename=filename)
        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(4, row_ct + 1):
            row = []
            for j in range(1, col_ct + 1):
                row.append(sh.cell(row=i, column=j).value)
            datalist.append(row)

        return datalist


class GradeSettingTest(unittest.TestCase):
    def __init__(self, test_method_name, test_data):
        super().__init__(test_method_name)
        self.test_data = test_data

    def setUp(self):
        self.driver = webdriver.Chrome(service=Service("./chromedriver"))
        self.driver.maximize_window()
        self.driver.get("https://school.moodledemo.net/course/modedit.php?update=852&return=1")
        Utils.login(self.driver)

    def test_grade_setting_for_assignment(self):
        driver = self.driver

        driver.get("https://school.moodledemo.net/course/modedit.php?update=852&return=1")

        test_id, grade_maximum, grade_to_pass, expected_error, expected_url = self.test_data

        wait = WebDriverWait(driver, 9999)
        openGradeSetting = wait.until(EC.element_to_be_clickable((By.ID, "collapseElement-7")))
        openGradeSetting.click()

        # Wait for the element to become visible and enabled
        wait = WebDriverWait(driver, 9999)
        maximum_grade = wait.until(EC.element_to_be_clickable((By.ID, "id_grade_modgrade_point")))

        maximum_grade.clear()
        if grade_maximum is not None:
            maximum_grade.send_keys(grade_maximum)

        grade_pass = driver.find_element(by=By.ID, value="id_gradepass")
        grade_pass.clear()
        if grade_to_pass is not None:
            grade_pass.send_keys(grade_to_pass)

        save_and_display = wait.until(EC.element_to_be_clickable((By.ID, "id_submitbutton")))
        save_and_display.click()

        # Check if grade setting was successful
        if expected_error is None:
            self.assertEqual(driver.current_url, expected_url)
        else:
            error_list = expected_error.split("; ")

            # Check if error message is displayed
            try:
                errors = driver.find_elements(
                    by=By.XPATH, value="//div[contains(@class, 'form-control-feedback')][normalize-space()]"
                )
            except NoSuchElementException:
                self.fail(f"Test {test_id} of test_grade_setting_for_assignment failed")

            if len(errors) != len(error_list):
                self.fail(f"Test {test_id} of test_grade_setting_for_assignment failed")

            for error in errors:
                if error.text not in error_list:
                    self.fail(f"Test {test_id} of test_grade_setting_for_assignment failed")

    def tearDown(self):
        self.driver.close()


if __name__ == "__main__":
    # Read test data from Excel file
    testcases = Utils.read_data_from_excel("./testcase.xlsx", "testdata")

    # Create a test suite and add the GradeSettingTest class
    suite = unittest.TestSuite()

    # Loop through test cases and add a test case for each one
    for test_data in testcases:
        test_case = GradeSettingTest("test_grade_setting_for_assignment", test_data=test_data)
        suite.addTest(test_case)

    # Run the test suite
    runner = unittest.TextTestRunner(verbosity=2)
    runner.run(suite)
